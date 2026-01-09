import pandas as pd
import numpy as np
import shutil
import os
import warnings
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

# Matikan warning agar terminal bersih
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

# ==============================================================================
# CONFIGURATION
# ==============================================================================
CONFIG = {
    "INPUT_FILE": "12 Lampiran Pendapatan Desember 2025.xlsx", 
    "FILE_REKAP": "01 251110 Validasi Kode Produk Rekap New.xlsx",
    "FILE_PELANGGAN": "12. BANGER-PELANGGAN-12122025.xlsx",
    "FILE_OPT": "12. BANGER-OPT-12122025.xlsx",
    "TEMPLATE_FILE": "11 BANGER NOVEMBER (2).xlsx", 
    "OUTPUT_FILE": "12 BANGER DESEMBER main.xlsx",
    "BULAN_LALU": "November",   
    "BULAN_INI": "Desember",   
    "DASHBOARD_HEADER_ROW": 2, 
    "DASHBOARD_DATA_START": 3
}

# === CUSTOM SORT ORDER untuk kodeMasterProduk (dari mentor) ===
# UPDATE: 61 produk (revisi dari lampiran konsol terbaru)
CUSTOM_KODE_PRODUK_ORDER = [
    756, 719, 731, 714, 464, 715, 737, 750, 725, 740, 747, 727, 739, 749, 748,
    721, 743, 745, 744, 735, 717, 738, 692, 673, 698, 693, 699, 697, 696, 694,
    695, 723, 746, 736, 713, 730, 728, 729, 755, 741, 724, 726, 757, 758, 732,
    273, 425, 435, 436, 734, 733, 722, 274, 708, 709, 710, 764, 768, 766, 767,
    487
]

# --- HELPERS ---
def super_clean(val):
    if pd.isna(val): return ""
    return str(val).strip().lower().replace('.0', '')

def safe_write(ws, row, col, value, fill=None, font=None):
    """Write to Excel cell with error handling (non-critical errors only)"""
    try:
        cell = ws.cell(row=row, column=col)
        if value is not None: cell.value = value
        if fill: cell.fill = fill
        if font: cell.font = font
    except Exception as e:
        print(f"⚠️  Warning: Gagal menulis cell ({row}, {col}): {e}")

def clean_header(header):
    """Normalize header untuk anti-typo mapping (lowercase, hapus symbol/spaces)"""
    import re
    if pd.isna(header): return ""
    return re.sub(r'[^a-z0-9]', '', str(header).lower())

def find_column_by_name(ws, col_name, header_row=3, fuzzy=True):
    """Cari index kolom berdasarkan nama header (anti-hardcode magic numbers)
    
    Args:
        ws: worksheet openpyxl
        col_name: nama kolom yang dicari (case-insensitive)
        header_row: baris header (default row 3)
        fuzzy: allow partial match (default True)
    
    Returns:
        int: column index (1-based) atau None jika tidak ketemu
    """
    col_clean = clean_header(col_name)
    for col_idx in range(1, ws.max_column + 1):
        cell_val = ws.cell(header_row, col_idx).value
        if not cell_val:
            continue
        cell_clean = clean_header(str(cell_val))
        if fuzzy:
            if col_clean in cell_clean or cell_clean in col_clean:
                return col_idx
        else:
            if col_clean == cell_clean:
                return col_idx
    return None

def validate_required_columns(df, required_cols, context="DataFrame"):
    """Validasi kolom wajib ada di DataFrame (fail fast dengan error jelas)
    
    Args:
        df: pandas DataFrame
        required_cols: list kolom yang wajib ada
        context: nama konteks untuk error message
    
    Raises:
        ValueError: jika ada kolom yang hilang
    """
    missing = [col for col in required_cols if col not in df.columns]
    if missing:
        raise ValueError(f"❌ {context}: Kolom wajib tidak ditemukan: {missing}\nKolom tersedia: {list(df.columns)[:10]}")

def custom_sort_by_kode_produk(df, kode_col):
    """Sort DataFrame berdasarkan custom order kodeMasterProduk dan DROP yang tidak ada di list
    
    Args:
        df: pandas DataFrame
        kode_col: nama kolom kode produk
    
    Returns:
        DataFrame yang sudah di-sort dan di-filter
    """
    if kode_col not in df.columns:
        print(f"     ⚠️  Warning: Kolom '{kode_col}' tidak ditemukan, skip custom sorting")
        return df
    
    # Buat set untuk cek keberadaan kode (lebih cepat dari list)
    valid_kode_set = set(CUSTOM_KODE_PRODUK_ORDER)
    
    # Filter: HANYA ambil data yang kode produknya ada di list
    before_filter = len(df)
    df_filtered = df[df[kode_col].apply(
        lambda x: int(x) in valid_kode_set if pd.notna(x) and str(x).replace('.','').replace('-','').isdigit() else False
    )].copy()
    after_filter = len(df_filtered)
    
    dropped_count = before_filter - after_filter
    if dropped_count > 0:
        print(f"     🗑️  Dropped {dropped_count} rows (kode produk tidak ada di list 63 kode)")
    
    # Buat mapping kode → urutan (756=0, 719=1, 731=2, ...)
    kode_order_map = {kode: idx for idx, kode in enumerate(CUSTOM_KODE_PRODUK_ORDER)}
    
    # Tambahkan kolom temporary untuk sort order
    df_filtered['_sort_order'] = df_filtered[kode_col].apply(
        lambda x: kode_order_map.get(int(x) if pd.notna(x) else x, 9999)
    )
    
    # Sort by temporary column
    df_sorted = df_filtered.sort_values(by='_sort_order', ascending=True).drop(columns=['_sort_order'])
    
    # RESET INDEX supaya row numbering rapat (0, 1, 2, 3... tanpa gap)
    # Ini mencegah row kosong setelah filter + sort
    df_sorted = df_sorted.reset_index(drop=True)
    
    print(f"     ✅ Data sorted by '{kode_col}' (custom order: 756→719→731→...) - {len(df_sorted)} rows")
    return df_sorted

# ==============================================================================
# FASE 1: EXTRACT (GLOBAL SEARCH)
# ==============================================================================
# FASE 1: EXTRACT (GLOBAL SEARCH)
# ==============================================================================
def extract_data():
    print(f"🚀 [1/3] EXTRACT: Mapping Master Data...")
    
    # Validasi file exists
    if not os.path.exists(CONFIG["FILE_REKAP"]):
        raise FileNotFoundError(f"❌ File Rekap tidak ditemukan: {CONFIG['FILE_REKAP']}")
    if not os.path.exists(CONFIG["INPUT_FILE"]):
        raise FileNotFoundError(f"❌ File Input tidak ditemukan: {CONFIG['INPUT_FILE']}")
    
    try:
        df_pdf = pd.read_excel(CONFIG["FILE_REKAP"], sheet_name="ALL PRODUCT PDF", header=2)
    except ValueError as e:
        raise ValueError(f"❌ Sheet 'ALL PRODUCT PDF' tidak ditemukan di {CONFIG['FILE_REKAP']}. Error: {e}")
    
    df_pdf.columns = [str(c).strip() for c in df_pdf.columns]
    
    # Validasi kolom wajib
    validate_required_columns(df_pdf, ['ICON+ Product'], context="ALL PRODUCT PDF")
    
    # Mencari nama produk di semua level segmen
    potential_name_cols = ['Product Portofolio Segmen 1', 'Product Portofolio Segmen 2', 'Product Portofolio Segmen 3', 'SEGMEN']
    prod_to_kode = {}
    for _, row in df_pdf.iterrows():
        kode_icon = str(row['ICON+ Product']).strip().replace('.0', '')
        if not kode_icon or kode_icon == 'nan': continue
        for col in potential_name_cols:
            if col in df_pdf.columns:
                nama_raw = str(row[col]).strip()
                if nama_raw and nama_raw != 'nan':
                    prod_to_kode[super_clean(nama_raw)] = kode_icon

    # Fallback ke tab SAP jika ada yang belum tercover
    try:
        xls_rekap = pd.ExcelFile(CONFIG["FILE_REKAP"])
        sap_sheet = next((s for s in xls_rekap.sheet_names if "sap" in s.lower()), None)
        if sap_sheet:
            df_sap = pd.read_excel(CONFIG["FILE_REKAP"], sheet_name=sap_sheet, header=1)
            validate_required_columns(df_sap, ['Nama Produk', 'Kode di SAP'], context="SAP Sheet")
            for _, row in df_sap.iterrows():
                nama_sap = str(row['Nama Produk']).strip()
                kode_sap = str(row['Kode di SAP']).strip().replace('.0', '')
                if nama_sap != 'nan' and super_clean(nama_sap) not in prod_to_kode:
                    prod_to_kode[super_clean(nama_sap)] = kode_sap
    except Exception as e:
        print(f"⚠️  Warning: Gagal load SAP sheet (optional): {e}")

    # Detail Portofolio untuk tab Realisasi
    portfolio_cols = ['ICON+ Product', 'Business Portofolio Segment 0', 'Kode 0', 'Product Portofolio Segmen 1', 'Kode 1', 'Product Portofolio Segmen 2', 'Kode 2', 'Product Portofolio Segmen 3', 'Kode 3', 'SEGMEN']
    df_portfolio = df_pdf[[c for c in portfolio_cols if c in df_pdf.columns]].copy()
    df_portfolio['Join_Key'] = df_portfolio['ICON+ Product'].apply(super_clean)
    df_portfolio = df_portfolio.drop_duplicates(subset=['Join_Key'])

    # Baca Raw Konsol
    xls_raw = pd.ExcelFile(CONFIG["INPUT_FILE"])
    target_sheet = next((s for s in xls_raw.sheet_names if "Konsol" in s), None)
    
    if not target_sheet:
        raise ValueError(f"❌ Sheet 'Konsol' tidak ditemukan di {CONFIG['INPUT_FILE']}. Sheet tersedia: {xls_raw.sheet_names}")
    
    # Cari header row: cari row yang ada "Customer No" atau "Customer Name"
    # Handle 2 struktur: November (ada Row Labels) vs Desember (tanpa Row Labels)
    df_tmp = pd.read_excel(xls_raw, sheet_name=target_sheet, header=None, nrows=20)
    h_row = 0
    has_row_labels = False
    for i, r in df_tmp.iterrows():
        row_str = r.astype(str).str.lower().tolist()
        if any(k in row_str for k in ["customer no", "customer name", "customer number"]):
            h_row = i
            # Check apakah ada "Row Labels" di kolom pertama
            if "row labels" in str(r.iloc[0]).lower():
                has_row_labels = True
            break
    
    # BACA NAMA PRODUK dari row sebelum header (h_row - 1)
    # Row h_row-1 = nama produk lengkap, Row h_row = kode produk
    kode_to_nama_produk = {}
    if h_row > 0:
        row_nama = df_tmp.iloc[h_row - 1]  # Row nama produk
        row_kode = df_tmp.iloc[h_row]      # Row kode/header
        
        for col_idx in range(len(row_kode)):
            kode_val = row_kode.iloc[col_idx]
            nama_val = row_nama.iloc[col_idx]
            
            # Skip Customer No/Name columns
            if pd.notna(kode_val) and str(kode_val).lower() not in ['customer no', 'customer name', 'customer number', 'row labels']:
                # Kode biasanya angka (120, 121, dst)
                if isinstance(kode_val, (int, float)) or str(kode_val).isdigit():
                    kode_clean = str(int(float(kode_val)))
                    # Nama produk dari row di atas (bisa kosong kalau merged cell)
                    if pd.notna(nama_val) and str(nama_val) != 'nan':
                        kode_to_nama_produk[kode_clean] = str(nama_val).strip()
    
    print(f"   > Mapped {len(kode_to_nama_produk)} kode → nama produk (contoh: {list(kode_to_nama_produk.items())[:3]})")
    
    df_raw = pd.read_excel(xls_raw, sheet_name=target_sheet, header=h_row)
    
    # FIX: Rename kolom "Customer No" → "Customer Number" (ignore "Row Labels")
    rename_map = {}
    
    # Hapus kolom "Row Labels" jika ada (tidak dipakai)
    if 'Row Labels' in df_raw.columns:
        df_raw = df_raw.drop(columns=['Row Labels'])
        print(f"   > Dropped 'Row Labels' column (not needed)")
    
    # Cari dan rename Customer No
    if 'Customer No' in df_raw.columns:
        rename_map['Customer No'] = 'Customer Number'
    elif 'Customer NO' in df_raw.columns:
        rename_map['Customer NO'] = 'Customer Number'
    elif 'Customer Number' not in df_raw.columns:
        raise ValueError(f"❌ Error: Kolom 'Customer No' atau 'Customer NO' tidak ditemukan di sheet Konsol. Kolom tersedia: {list(df_raw.columns[:10])}")
    
    # Customer Name tidak perlu di-rename (sudah sesuai)
    if 'Customer Name' not in df_raw.columns:
        print(f"   ⚠️  Warning: Kolom 'Customer Name' tidak ditemukan, akan skip kolom ini")
    
    if rename_map:
        df_raw.rename(columns=rename_map, inplace=True)
        print(f"   > Column mapping: {rename_map}")
    
    # CRITICAL FIX: JANGAN pakai ffill! 
    # Hanya ambil rows yang ASLI punya Customer Number (bukan hasil merged cells)
    # Rows dengan Customer Number = NaN = bukan data customer (subtotal, header, dll)
    before_drop = len(df_raw)
    df_raw = df_raw[df_raw['Customer Number'].notna()].copy()
    after_drop = len(df_raw)
    
    if before_drop > after_drop:
        print(f"   > Dropped {before_drop - after_drop} rows tanpa Customer Number (subtotal/header/merged cells)")
    
    # DEFENSE LAYER 2: Filter Grand Total di SEMUA kolom (handle file Oktober yang broken)
    # Cek semua kolom (Row Labels, Customer Number, Customer Name, dll) untuk string "grand total"
    before_total_filter = len(df_raw)
    mask_grand_total = df_raw.apply(lambda row: row.astype(str).str.lower().str.contains('grand total', na=False).any(), axis=1)
    df_raw = df_raw[~mask_grand_total].copy()
    after_total_filter = len(df_raw)
    
    if before_total_filter > after_total_filter:
        print(f"   > Dropped {before_total_filter - after_total_filter} rows dengan 'Grand Total' (total/subtotal rows)")
    
    # DEFENSE LAYER 3: Filter pivot table summary rows (Digital Platform, PLN Group, Publik, Retail, dll)
    # Detect keywords yang menandakan ini bukan customer data
    # Check di Customer Name (bukan di semua kolom) untuk lebih presisi
    before_summary_filter = len(df_raw)
    
    # Filter berdasarkan Customer Number yang valid (harus angka murni, bukan text)
    # Dan Customer Name yang bukan summary row
    def is_valid_customer_row(row):
        cust_no = str(row['Customer Number']).strip().replace('.0', '').replace('.', '')
        cust_name = str(row.get('Customer Name', '')).lower().strip() if 'Customer Name' in row.index else ''
        
        # Check Customer Number: harus angka murni (bukan text kayak "Digital Platform")
        # Relaxed check: asal digit aja, tidak perlu cek prefix 20000
        if not cust_no.isdigit():
            return False
        
        # Check panjang minimal (customer number biasanya 8-10 digit)
        if len(cust_no) < 8:
            return False
        
        # Check Customer Name: filter ONLY exact summary rows (lebih ketat)
        # Gunakan exact match atau keyword yang sangat specific
        if cust_name in ['digital platform', 'pln group', 'publik', 'retail']:
            return False
        
        # Keyword yang pasti summary row (lebih aman)
        if any(keyword in cust_name for keyword in ['sum of', 'persentase', 'percentage', 
                                                     'row labels', 'column labels', 'subtotal', 'total amount']):
            return False
        
        return True
    
    df_raw = df_raw[df_raw.apply(is_valid_customer_row, axis=1)].copy()
    after_summary_filter = len(df_raw)
    
    if before_summary_filter > after_summary_filter:
        print(f"   > Dropped {before_summary_filter - after_summary_filter} rows pivot summary (Digital Platform, PLN Group, Publik, Retail, dll)")
    
    # DEBUG: Print jumlah data yang tersisa
    print(f"   > Final data count: {len(df_raw)} rows (customer data yang valid)")

    return prod_to_kode, df_portfolio, df_raw, kode_to_nama_produk

# ==============================================================================
# FASE 2: TRANSFORM
# ==============================================================================
def transform_data(prod_to_kode, df_portfolio, df_raw, kode_to_nama_produk):
    print("⚙️ [2/3] TRANSFORM: Unpivoting & Matching...")
    
    # Validasi input tidak kosong
    if df_raw.empty:
        raise ValueError("❌ Data input kosong! Tidak ada data untuk diproses.")
    if not prod_to_kode:
        print("⚠️  WARNING: Mapping produk kosong! Semua produk akan ter-map ke NULL.")
    
    cols = df_raw.columns.tolist()
    
    # Cari kolom Customer Number (sudah di-rename dari "Customer NO" di extract_data)
    if 'Customer Number' not in cols:
        raise ValueError(f"❌ Kolom 'Customer Number' tidak ditemukan! Kolom tersedia: {cols[:5]}...")
    
    # Tentukan kolom ID yang akan di-unpivot (Customer Number + Customer Name)
    id_cols = ['Customer Number']
    if 'Customer Name' in cols:
        id_cols.append('Customer Name')
    
    # Filter produk columns: skip Customer, Total, Unnamed, dan Segmen
    prod_cols = [c for c in cols if c not in id_cols 
                 and "total" not in str(c).lower() 
                 and "unnamed" not in str(c).lower()
                 and str(c).lower().strip() != "segmen"]
    
    print(f"   > Customer Columns: {id_cols}")
    print(f"   > Product Columns: {len(prod_cols)} produk (contoh: {prod_cols[:3]}...)")
    
    df_melt = pd.melt(df_raw, id_vars=id_cols, value_vars=prod_cols, var_name='Produk_Raw', value_name='Value')
    df_melt['Value'] = pd.to_numeric(df_melt['Value'], errors='coerce').fillna(0)
    df_melt = df_melt[df_melt['Value'] != 0].copy()
    
    # Mapping produk: cek apakah Produk_Raw sudah kode (angka) atau nama (text)
    def get_kode_produk(prod_raw):
        # Handle numeric types (int, float)
        if isinstance(prod_raw, (int, float)):
            return str(int(prod_raw))
        
        prod_str = str(prod_raw).strip()
        
        # Cek apakah string numeric (bisa ada .0)
        prod_clean = prod_str.replace('.0', '').replace('.', '')
        if prod_clean.isdigit() and len(prod_clean) >= 3:
            return prod_clean
        
        # Kalau bukan angka, cari di mapping nama → kode
        return prod_to_kode.get(super_clean(prod_raw), None)
    
    df_melt['Kode Produk'] = df_melt['Produk_Raw'].apply(get_kode_produk)
    df_melt['Join_Key'] = df_melt['Kode Produk'].apply(super_clean)
    
    # Map kode produk ke nama lengkap (dari row nama produk di Konsol)
    # Kalau kode produk adalah angka (121, 122), cari nama lengkapnya
    def get_nama_produk_lengkap(row):
        prod_raw = row['Produk_Raw']
        kode = row['Kode Produk']
        
        # Kalau Produk_Raw sudah nama (string panjang), pakai itu
        if isinstance(prod_raw, str) and len(prod_raw) > 10:
            return prod_raw
        
        # Kalau kode produk ada mapping ke nama, pakai nama lengkap
        if pd.notna(kode) and str(kode) in kode_to_nama_produk:
            return kode_to_nama_produk[str(kode)]
        
        # Fallback: pakai Produk_Raw apa adanya
        return str(prod_raw)
    
    df_melt['Nama Produk Lengkap'] = df_melt.apply(get_nama_produk_lengkap, axis=1)
    
    # DEBUG: Cek berapa produk yang tidak ada di portfolio
    unique_kode = df_melt['Kode Produk'].unique()
    portfolio_keys = df_portfolio['Join_Key'].unique()
    missing_kode = [k for k in unique_kode if super_clean(k) not in portfolio_keys]
    if missing_kode:
        print(f"   ⚠️  WARNING: {len(missing_kode)} kode produk tidak ditemukan di ALL PRODUCT PDF:")
        print(f"      Contoh: {missing_kode[:5]}")
    
    # Merge dengan portfolio data
    df_final = df_melt.merge(df_portfolio, on='Join_Key', how='left')
    
    # Drop kolom duplikat ICON+ Product dari portfolio (karena sudah ada Kode Produk)
    if 'ICON+ Product' in df_final.columns:
        df_final = df_final.drop(columns=['ICON+ Product'])
    
    # Drop Produk_Raw, pakai Nama Produk Lengkap sebagai Produk/Layanan
    if 'Produk_Raw' in df_final.columns:
        df_final = df_final.drop(columns=['Produk_Raw'])
    df_final.rename(columns={'Nama Produk Lengkap': 'Produk/Layanan'}, inplace=True)
    
    # Urutan kolom final: Customer Number, Customer Name (jika ada), Kode Produk, dst
    final_cols = ['Customer Number']
    if 'Customer Name' in df_final.columns:
        final_cols.append('Customer Name')
    final_cols.extend(['Kode Produk', 'Produk/Layanan', 'Value', 'Business Portofolio Segment 0', 'Kode 0', 'Product Portofolio Segmen 1', 'Kode 1', 'Product Portofolio Segmen 2', 'Kode 2', 'Product Portofolio Segmen 3', 'Kode 3', 'SEGMEN'])
    
    for c in final_cols:
        if c not in df_final.columns: df_final[c] = ""
    
    return df_final[final_cols]


# ==============================================================================
# FASE 3: LOAD (DINAMIS S-R LOGIC)
# ==============================================================================
def load_data(df_final):
    print("💾 [3/3] LOAD: Update Dashboard dengan Logika Dinamis S-R...")
    
    # 1. AMBIL SALDO OKTOBER (LALU) SEBAGAI ANGKA
    wb_master = load_workbook(CONFIG["TEMPLATE_FILE"], data_only=True, keep_vba=False)
    ws_dash_m = wb_master["Dashboard"]
    
    # Cari indeks kolom Oktober (Bulan Lalu)
    col_lalu_idx = next((c.column for c in ws_dash_m[CONFIG["DASHBOARD_HEADER_ROW"]] 
                        if CONFIG["BULAN_LALU"].lower() in str(c.value).lower() and "kumulatif" in str(c.value).lower()), None)
    
    master_lalu_vals = {r: ws_dash_m.cell(r, col_lalu_idx).value or 0 
                        for r in range(CONFIG["DASHBOARD_DATA_START"], ws_dash_m.max_row + 1)} if col_lalu_idx else {}
    wb_master.close()

    # 2. PROSES FILE OUTPUT
    shutil.copyfile(CONFIG["TEMPLATE_FILE"], CONFIG["OUTPUT_FILE"])
    wb = load_workbook(CONFIG["OUTPUT_FILE"])
    
    # Update Tab Realisasi (Rename & Overwrite)
    new_sheet = f"Realisasi {CONFIG['BULAN_INI']}"
    old_sheet = f"Realisasi {CONFIG['BULAN_LALU']}"
    if old_sheet in wb.sheetnames:
        ws_real = wb[old_sheet]; ws_real.title = new_sheet
    else:
        ws_real = wb.create_sheet(new_sheet)
    
    if ws_real.max_row > 1: ws_real.delete_rows(2, ws_real.max_row)
    
    # === SORT df_final by 61 kode produk sebelum tulis ke Realisasi ===
    print(f"   > Sorting Realisasi data by 61 kode produk...")
    df_final_sorted = custom_sort_by_kode_produk(df_final.copy(), 'Kode Produk')
    
    for c, name in enumerate(df_final_sorted.columns, 1): ws_real.cell(1, c).value = name
    for r_idx, row in enumerate(dataframe_to_rows(df_final_sorted, index=False, header=False), 2):
        for c_idx, val in enumerate(row, 1): ws_real.cell(r_idx, c_idx).value = val
    
    print(f"     ✅ Realisasi {CONFIG['BULAN_INI']} updated: {len(df_final_sorted)} rows (sorted by 61 kode)")

    # 3. UPDATE DASHBOARD (LOGIKA DINAMIS S-R)
    ws_dash = wb["Dashboard"]
    target_month = CONFIG["BULAN_INI"].lower()
    
    # Mapping bulan ke index (untuk hitung pembagi prognosa)
    BULAN_MAP = {"januari": 1, "februari": 2, "maret": 3, "april": 4, "mei": 5, "juni": 6,
                 "juli": 7, "agustus": 8, "september": 9, "oktober": 10, "november": 11, "desember": 12}
    bulan_index = BULAN_MAP.get(target_month, 1)
    
    # === HARDCODE KOLOM DASHBOARD (TEMPLATE-SPECIFIC BUSINESS LOGIC) ===
    # NOTE: Kolom ini di-hardcode karena merupakan struktur template yang sudah FIX
    # dan memiliki logika bisnis spesifik (S-R calculation, prognosa)
    # Jika template berubah struktur, update konstanta di bawah:
    COL_SA_JAN = 23    # W - SA Januari
    COL_SA_DES = 34    # AH - SA Desember (Prognosa Bulanan)
    COL_KUM_JAN = 10   # J - Kumulatif Januari
    COL_KUM_DES = 21   # U - Kumulatif Desember (Prognosa Kumulatif)
    COL_SISA = 22      # V - Sisa Perhitungan CO NR
    
    # === MAPPING KOLOM DATA PELANGGAN (untuk SUMIF Prognosa) ===
    KOLOM_CARRY_OVER_START = 51   # AY = Januari Carry Over
    KOLOM_NEW_REVENUE_START = 63  # BK = Januari New Revenue
    
    # Konversi ke letter
    LET_SA_JAN = get_column_letter(COL_SA_JAN)    # W
    LET_SA_DES = get_column_letter(COL_SA_DES)    # AH
    LET_KUM_DES = get_column_letter(COL_KUM_DES)  # U
    LET_SISA = get_column_letter(COL_SISA)        # V
    
    # Cari kolom bulan lalu untuk ambil nilai
    col_lalu_idx = next((c.column for c in ws_dash[CONFIG["DASHBOARD_HEADER_ROW"]] 
                        if CONFIG["BULAN_LALU"].lower() in str(c.value).lower() and "kumulatif" in str(c.value).lower()), None)
    
    # Cari semua kolom yang ada nama bulan ini (untuk update data)
    cols_found = [c.column for c in ws_dash[CONFIG["DASHBOARD_HEADER_ROW"]] if target_month in str(c.value).lower()]
    
    if cols_found and col_lalu_idx:
        col_kum = cols_found[0]  # Kumulatif (Misal Kolom S)
        col_sa = cols_found[1] if len(cols_found) > 1 else col_kum + 1 # SA (Bulan Berjalan)
        
        let_kum = get_column_letter(col_kum)
        let_sa = get_column_letter(col_sa)
        let_lalu = get_column_letter(col_lalu_idx)
        
        # HARDCODE GRAND TOTAL ROW = 64 (template stabil)
        grand_total_row = 64
        print(f"   > Grand Total row: {grand_total_row} (hardcoded)")

        print(f"   > Logika Dinamis: SA {let_sa} = {let_kum} - {let_lalu}")
        print(f"   > Prognosa: AH (Desember) = Rata-rata W (Jan) s/d {let_sa} / {bulan_index}")
        print(f"   > Sisa Perhitungan V = {'AH' if bulan_index == 11 else f'SUM kolom prognosa sisa'}")

        for r in range(CONFIG["DASHBOARD_DATA_START"], ws_dash.max_row + 1):
            cell_A = ws_dash.cell(r, 1).value
            if not cell_A: continue
            
            # DETEKSI: Apakah ini FIRST DATA ROW atau GRAND TOTAL ROW
            is_first_row = (r == CONFIG["DASHBOARD_DATA_START"])
            # Grand Total detection: check exact row number (sudah di-detect sebelumnya)
            is_grand_total = (r == grand_total_row)
            
            # FORCE CLEAR kolom yang akan di-update
            ws_dash.cell(r, col_kum).value = None
            ws_dash.cell(r, col_sa).value = None
            
            # Clear prognosa & sisa
            if is_first_row:
                # First row: Clear semua kolom prognosa (AH, U, V)
                ws_dash.cell(r, COL_SA_DES).value = None   # AH - Prognosa Bulanan
                ws_dash.cell(r, COL_KUM_DES).value = None  # U - Prognosa Kumulatif
                ws_dash.cell(r, COL_SISA).value = None     # V - Sisa Perhitungan
            else:
                # Row lain: Clear kolom Sisa Perhitungan saja (prognosa biarkan template)
                ws_dash.cell(r, COL_SISA).value = None
                
                # KHUSUS DESEMBER: Clear SEMUA kolom prognosa (tidak ada prognosa lagi)
                if bulan_index == 12:
                    # Clear prognosa bulan berikutnya (tidak ada karena sudah Des)
                    # Loop dari bulan setelah bulan ini (Des+1 = tidak ada) s/d Des (12)
                    for prog_month_idx in range(bulan_index + 1, 13):  # Tidak akan loop (13, 13)
                        col_prog_sa = COL_SA_JAN + prog_month_idx - 1
                        ws_dash.cell(r, col_prog_sa).value = None
            
            # Tempel Nilai Bulan Lalu (Hard Value)
            ws_dash.cell(r, col_lalu_idx).value = master_lalu_vals.get(r, 0)
            
            # ========== CASE KHUSUS: JANUARI ==========
            if bulan_index == 1:
                # JANUARI: Tidak ada bulan lalu, semua prognosa = copy Januari
                
                if "total" in str(cell_A).lower():
                    # Grand Total: SUM semua kolom
                    ws_dash.cell(r, col_kum).value = f"=SUM({let_kum}{CONFIG['DASHBOARD_DATA_START']}:{let_kum}{r-1})"
                    ws_dash.cell(r, col_sa).value = f"=SUM({let_sa}{CONFIG['DASHBOARD_DATA_START']}:{let_sa}{r-1})"
                    
                    # Prognosa Feb-Des: SUM vertikal
                    for m_idx in range(2, 13):  # Feb=2, Mar=3, ..., Des=12
                        col_prog_sa = COL_SA_JAN + m_idx - 1
                        col_prog_kum = COL_KUM_JAN + m_idx - 1
                        let_prog_sa_temp = get_column_letter(col_prog_sa)
                        let_prog_kum_temp = get_column_letter(col_prog_kum)
                        ws_dash.cell(r, col_prog_sa).value = f"=SUM({let_prog_sa_temp}{CONFIG['DASHBOARD_DATA_START']}:{let_prog_sa_temp}{r-1})"
                        ws_dash.cell(r, col_prog_kum).value = f"=SUM({let_prog_kum_temp}{CONFIG['DASHBOARD_DATA_START']}:{let_prog_kum_temp}{r-1})"
                    
                    ws_dash.cell(r, COL_SISA).value = f"=SUM({LET_SISA}{CONFIG['DASHBOARD_DATA_START']}:{LET_SISA}{r-1})"
                
                else:
                    # Data Row
                    # 1. Kumulatif Januari = SUMIF
                    formula_sumif = f"SUMIF('{new_sheet}'!$C:$C, A{r}, '{new_sheet}'!$E:$E)"
                    ws_dash.cell(r, col_kum).value = f"={formula_sumif}"
                    
                    # 2. SA Januari = Kumulatif Januari (tidak ada bulan lalu)
                    ws_dash.cell(r, col_sa).value = f"={let_kum}{r}"
                    
                    if is_first_row:
                        # === FIRST ROW: Prognosa = Copy nilai Januari ===
                        # Loop Feb-Des (kolom AB s/d AH untuk SA, K s/d U untuk Kumulatif)
                        for m_idx in range(2, 13):  # Feb=2, ..., Des=12
                            col_prog_sa = COL_SA_JAN + m_idx - 1      # AB, AC, AD, ..., AH
                            col_prog_kum = COL_KUM_JAN + m_idx - 1    # K, L, M, ..., U
                            
                            # SA Prognosa = Copy SA Januari
                            ws_dash.cell(r, col_prog_sa).value = f"={LET_SA_JAN}{r}"
                            
                            # Kumulatif Prognosa = Kumulatif sebelumnya + SA Prognosa
                            if m_idx == 2:  # Februari
                                ws_dash.cell(r, col_prog_kum).value = f"={let_kum}{r}+{get_column_letter(col_prog_sa)}{r}"
                            else:
                                col_prev_kum = col_prog_kum - 1
                                ws_dash.cell(r, col_prog_kum).value = f"={get_column_letter(col_prev_kum)}{r}+{get_column_letter(col_prog_sa)}{r}"
                    
                    # Sisa Perhitungan: SUM Feb-Des (11 bulan)
                    col_feb_sa = COL_SA_JAN + 1  # AB = SA Februari
                    let_feb_sa = get_column_letter(col_feb_sa)
                    ws_dash.cell(r, COL_SISA).value = f"=SUM({let_feb_sa}{r}:{LET_SA_DES}{r})"
            
            # ========== CASE NORMAL: FEB-DES ==========
            elif is_grand_total:
                # GRAND TOTAL ROW: SUM vertikal semua data row (BUKAN SUMIF!)
                ws_dash.cell(r, col_kum).value = f"=SUM({let_kum}{CONFIG['DASHBOARD_DATA_START']}:{let_kum}{r-1})"
                ws_dash.cell(r, col_sa).value = f"=SUM({let_sa}{CONFIG['DASHBOARD_DATA_START']}:{let_sa}{r-1})"
                
                # Prognosa Grand Total (SUM vertikal) - HARDCODE
                if bulan_index < 12:
                    ws_dash.cell(r, COL_SA_DES).value = f"=SUM({LET_SA_DES}{CONFIG['DASHBOARD_DATA_START']}:{LET_SA_DES}{r-1})"
                    ws_dash.cell(r, COL_KUM_DES).value = f"=SUM({LET_KUM_DES}{CONFIG['DASHBOARD_DATA_START']}:{LET_KUM_DES}{r-1})"
                    ws_dash.cell(r, COL_SISA).value = f"=SUM({LET_SISA}{CONFIG['DASHBOARD_DATA_START']}:{LET_SISA}{r-1})"
                elif bulan_index == 12:
                    # Desember: Grand Total juga SUM vertikal
                    ws_dash.cell(r, COL_SA_DES).value = f"=SUM({LET_SA_DES}{CONFIG['DASHBOARD_DATA_START']}:{LET_SA_DES}{r-1})"
                    ws_dash.cell(r, COL_KUM_DES).value = f"=SUM({LET_KUM_DES}{CONFIG['DASHBOARD_DATA_START']}:{LET_KUM_DES}{r-1})"
                    ws_dash.cell(r, COL_SISA).value = f"=SUM({LET_SISA}{CONFIG['DASHBOARD_DATA_START']}:{LET_SISA}{r-1})"
            else:
                # Formula Kumulatif: HANYA SUMIF (tidak pakai tambahan)
                formula_sumif = f"SUMIF('{new_sheet}'!$C:$C, A{r}, '{new_sheet}'!$E:$E)"
                ws_dash.cell(r, col_kum).value = f"={formula_sumif}"
                
                # Formula SA Dinamis: Kumulatif Sekarang - Kumulatif Lalu
                ws_dash.cell(r, col_sa).value = f"={let_kum}{r} - {let_lalu}{r}"
                
                # PROGNOSA & SISA PERHITUNGAN
                # KHUSUS: Kalau bulan upload = Desember, maka kolom U/AH (Prognosa Desember) jadi REALISASI (SUMIF)
                # Kalau bulan upload < Desember, kolom U/AH tetap PROGNOSA (rata-rata atau SUMIF Data Pelanggan)
                
                if bulan_index == 12:
                    # DESEMBER: Kolom U/AH adalah REALISASI (bukan prognosa lagi)
                    # SEMUA ROW: SUMIF dari Realisasi Desember
                    
                    # Kolom U (Kumulatif Desember) = SUMIF Realisasi Desember
                    formula_sumif_des = f"SUMIF('{new_sheet}'!$C:$C, A{r}, '{new_sheet}'!$E:$E)"
                    ws_dash.cell(r, COL_KUM_DES).value = f"={formula_sumif_des}"
                    
                    # Kolom AH (SA Desember) = Kumulatif Des - Kumulatif Nov
                    ws_dash.cell(r, COL_SA_DES).value = f"={LET_KUM_DES}{r}-{let_lalu}{r}"
                    
                    # Sisa Perhitungan = 0 (tidak ada bulan tersisa)
                    ws_dash.cell(r, COL_SISA).value = 0
                elif bulan_index < 12:
                    sisa_bulan = 12 - bulan_index
                    
                    if is_first_row:
                        # === FIRST ROW: Rata-rata untuk prognosa (baseline) ===
                        # 1. Prognosa BULANAN Desember (AH) = Rata-rata SA Jan s/d Bulan Ini
                        ws_dash.cell(r, COL_SA_DES).value = f"=SUM({LET_SA_JAN}{r}:{let_sa}{r})/{bulan_index}"
                        
                        # 2. Prognosa KUMULATIF Desember (U) = Kumulatif Bulan Ini + Prognosa Bulanan
                        ws_dash.cell(r, COL_KUM_DES).value = f"={let_kum}{r}+{LET_SA_DES}{r}"
                        
                        # 3. Loop Prognosa bulan-bulan setelah bulan upload (Nov+1 s/d Des-1)
                        # Upload Oktober → Loop Nov (bulan_index+1=11) s/d Nov (Des-1=11) → HANYA November
                        # Upload September → Loop Okt (10) s/d Nov (11) → Oktober & November
                        for prog_month_idx in range(bulan_index + 1, 12):  # 11 s/d 11 (November saja)
                            # Kolom SA Prognosa bulan ini (misal AB=Feb, AC=Mar, ..., AG=Nov)
                            col_prog_sa = COL_SA_JAN + prog_month_idx - 1
                            # Kolom Kumulatif Prognosa bulan ini (misal K=Feb, L=Mar, ..., T=Nov)
                            col_prog_kum = COL_KUM_JAN + prog_month_idx - 1
                            
                            # SA Prognosa = Copy dari Prognosa Desember (AH) - rata-rata yang sama
                            ws_dash.cell(r, col_prog_sa).value = f"={LET_SA_DES}{r}"
                            
                            # Kumulatif Prognosa = Kumulatif bulan sebelumnya + SA Prognosa bulan ini
                            col_prev_kum = col_prog_kum - 1
                            let_prev_kum = get_column_letter(col_prev_kum)
                            let_prog_sa = get_column_letter(col_prog_sa)
                            ws_dash.cell(r, col_prog_kum).value = f"={let_prev_kum}{r}+{let_prog_sa}{r}"
                    
                    else:
                        # === ROW 4-60: SUMIF dari Data Pelanggan ===
                        # Loop untuk semua bulan prognosa (bulan_index+1 s/d 12)
                        # Upload Oktober (bulan_index=10) → Loop Nov(11) & Des(12)
                        for prog_month_idx in range(bulan_index + 1, 13):  # 11, 12 (November, Desember)
                            # Hitung kolom Carry Over & New Revenue di Data Pelanggan
                            col_co = KOLOM_CARRY_OVER_START + (prog_month_idx - 1)  # AY+10=BI (Nov), AY+11=BJ (Des)
                            col_nr = KOLOM_NEW_REVENUE_START + (prog_month_idx - 1) # BK+10=BU (Nov), BK+11=BV (Des)
                            
                            let_co = get_column_letter(col_co)
                            let_nr = get_column_letter(col_nr)
                            
                            # Kolom SA Prognosa di Dashboard
                            col_prog_sa = COL_SA_JAN + prog_month_idx - 1
                            # Kolom Kumulatif Prognosa di Dashboard
                            col_prog_kum = COL_KUM_JAN + prog_month_idx - 1
                            
                            # Formula SUMIF: Carry Over + New Revenue
                            formula_prog = f"=SUMIF('Data Pelanggan'!$AO:$AO, Dashboard!A{r}, 'Data Pelanggan'!${let_co}:${let_co})" \
                                          f"+SUMIF('Data Pelanggan'!$AO:$AO, Dashboard!A{r}, 'Data Pelanggan'!${let_nr}:${let_nr})"
                            
                            ws_dash.cell(r, col_prog_sa).value = formula_prog
                            
                            # Kumulatif Prognosa = Kumulatif bulan sebelumnya + SA Prognosa bulan ini
                            col_prev_kum = col_prog_kum - 1
                            let_prev_kum = get_column_letter(col_prev_kum)
                            let_prog_sa = get_column_letter(col_prog_sa)
                            ws_dash.cell(r, col_prog_kum).value = f"={let_prev_kum}{r}+{let_prog_sa}{r}"

                    
                    # 3. Sisa Perhitungan (V) - SEMUA ROW
                    if sisa_bulan == 1:
                        # Sisa 1 bulan (Nov → Des aja)
                        ws_dash.cell(r, COL_SISA).value = f"={LET_SA_DES}{r}"
                    else:
                        # Sisa > 1 bulan: SUM dari prognosa berikutnya s/d Desember
                        # Hardcode: AH=Desember, mundur sesuai sisa_bulan
                        start_col = COL_SA_DES - sisa_bulan + 1
                        let_start = get_column_letter(start_col)
                        ws_dash.cell(r, COL_SISA).value = f"=SUM({let_start}{r}:{LET_SA_DES}{r})"

    # ========== UPDATE TAB SUMMARY (DINAMIS) ==========
    col_kum_bulan_ini = COL_KUM_JAN + bulan_index - 1  # Kumulatif bulan ini untuk Summary
    col_sa_bulan_ini = COL_SA_JAN + bulan_index - 1    # Stand Alone bulan ini untuk Summary
    
    # Grand Total row sudah di-set sebelumnya (hardcoded = 64)
    
    if "Summary" in wb.sheetnames:
        ws_summary = wb["Summary"]
        
        # Kolom Kumulatif & Stand Alone bulan ini
        let_kum_summary = get_column_letter(col_kum_bulan_ini)
        let_sa_summary = get_column_letter(col_sa_bulan_ini)
        
        # Update header C2 saja (B2 biarkan dari template - "Target Desember" tanpa rumus)
        ws_summary.cell(2, 3).value = f"Realisasi {CONFIG['BULAN_INI']}"
        
        # Update data row 3 (bukan row 9 yang itu label)
        # B3: Target Desember - BIARKAN dari template (tidak ada rumus, angka statis)
        # C3: Realisasi bulan ini → Dashboard kolom KUMULATIF bulan ini Grand Total
        ws_summary.cell(3, 3).value = f"=Dashboard!{let_kum_summary}{grand_total_row}"
        print(f"   > Summary C3 (Realisasi {CONFIG['BULAN_INI']}): =Dashboard!{let_kum_summary}{grand_total_row}")
    
    # ========== UPDATE SHEET DATA PELANGGAN & OPT ==========
    update_sheet_pelanggan(wb, bulan_index)
    update_sheet_opt(wb, bulan_index)
    
    # ========== UPDATE RUMUS SUMIF DINAMIS DI DASHBOARD ==========
    update_dashboard_sumif_formulas(wb)
    
    wb.save(CONFIG["OUTPUT_FILE"])
    print(f"✅ BERHASIL! Dashboard + Summary + Data Pelanggan + OPT dinamis untuk {CONFIG['BULAN_INI'].upper()}!")

# ==============================================================================
# UPDATE SHEET DATA PELANGGAN
# ==============================================================================
def update_sheet_pelanggan(wb, bulan_index):
    """Update sheet Data Pelanggan dengan mapping kolom anti-typo dan formula dinamis"""
    print(f"   > Updating sheet 'Data Pelanggan'...")
    
    if "Data Pelanggan" not in wb.sheetnames:
        print(f"     ⚠️ Sheet 'Data Pelanggan' tidak ditemukan!")
        return
    
    ws = wb["Data Pelanggan"]
    
    # Baca header template dari Row 3
    template_headers = []
    for col in range(1, 100):  # Scan sampai kolom 100
        val = ws.cell(3, col).value
        if val:
            template_headers.append((col, str(val).strip()))
    
    # Baca data baru dari file raw
    try:
        df_raw = pd.read_excel(CONFIG["FILE_PELANGGAN"], header=0)
    except FileNotFoundError:
        print(f"     ⚠️ File tidak ditemukan: {CONFIG['FILE_PELANGGAN']}")
        return
    except Exception as e:
        print(f"     ❌ Error membaca file Pelanggan: {e}")
        return
    
    # Mapping kolom raw ke template (anti-typo)
    raw_to_template = {}
    for raw_col in df_raw.columns:
        raw_clean = clean_header(raw_col)
        for template_col_idx, template_col_name in template_headers:
            template_clean = clean_header(template_col_name)
            if raw_clean == template_clean:
                raw_to_template[raw_col] = template_col_idx
                break
    
    # === SPECIAL MAPPINGS (kolom yang beda nama antara template vs raw) ===
    # Nama Pelanggan (template) ← namaPerusahaan (raw)
    if 'namaPerusahaan' in df_raw.columns:
        for idx, (col_idx, col_name) in enumerate(template_headers):
            if clean_header(col_name) == clean_header('Nama Pelanggan'):
                raw_to_template['namaPerusahaan'] = col_idx
                break
    
    # DEBUG: Print mapping untuk cek
    print(f"     DEBUG: {len(raw_to_template)} kolom ter-mapping dari {len(df_raw.columns)} kolom raw")
    print(f"     DEBUG: Raw columns: {list(df_raw.columns[:5])}...")
    
    # === SORT BY kodeMasterProduk (CUSTOM ORDER dari mentor) ===
    kode_produk_col = None
    for raw_col in df_raw.columns:
        if clean_header(raw_col) == clean_header('kodeMasterProduk'):
            kode_produk_col = raw_col
            break
    
    if kode_produk_col:
        df_raw = custom_sort_by_kode_produk(df_raw, kode_produk_col)
    else:
        print(f"     ⚠️  Warning: Kolom 'kodeMasterProduk' tidak ditemukan, skip sorting")
    
    # === UPDATE BULAN BERJALAN (DINAMIS: scan row 1-5) ===
    # Cari label "Bulan Berjalan" di row manapun (1-5), tulis nilai di kolom sebelahnya di row yang sama
    col_label = None
    row_label = None
    for row_idx in range(1, 6):  # Scan row 1-5
        for col_idx in range(1, 100):
            cell_val = ws.cell(row_idx, col_idx).value
            if cell_val and "bulan berjalan" in str(cell_val).lower():
                col_label = col_idx
                row_label = row_idx
                break
        if col_label:
            break
    
    if col_label and row_label:
        # Tulis nilai di kolom sebelah kanan (col_label + 1) di row yang sama
        col_value = col_label + 1
        ws.cell(row_label, col_value).value = bulan_index
        print(f"     ✅ Update Bulan Berjalan: Row {row_label}, {get_column_letter(col_label)} (label) → {get_column_letter(col_value)} (nilai={bulan_index})")
    else:
        # Fallback ke hardcoded row 2, col 48
        print("     ⚠️  Warning: Kolom 'Bulan Berjalan' tidak ditemukan, gunakan fallback row 2, col 48")
        ws.cell(2, 48).value = bulan_index
    
    # === COPY TEMPLATE ROW 4 FORMULAS ===
    # Simpan formula dari row 4 template untuk di-copy ke row baru
    import re
    template_formulas = {}
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(4, col)
        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
            template_formulas[col] = cell.value
    
    # === HAPUS DATA LAMA (dari row 4 ke bawah) ===
    if ws.max_row > 3:
        ws.delete_rows(4, ws.max_row - 3)
    
    # === TULIS DATA BARU (dari row 4) ===
    for row_idx, row_data in df_raw.iterrows():
        excel_row = row_idx + 4  # Start dari row 4
        
        # 1. TULIS DATA RAW DULU (prioritas tertinggi)
        for raw_col, template_col_idx in raw_to_template.items():
            val = row_data[raw_col]
            if pd.notna(val):
                # Clean invalid date strings (--/--/--, --, dll) untuk mencegah #VALUE!
                str_val = str(val).strip()
                if str_val in ['--/--/--', '--', '---', 'nan', 'NaT', 'None']:
                    # Skip: biarkan cell kosong
                    continue
                ws.cell(excel_row, template_col_idx).value = val
        
        # 2. Copy formulas dari template HANYA untuk kolom yang TIDAK ada di raw data
        for col_idx, formula_template in template_formulas.items():
            # Skip kalau kolom ini sudah ada data raw
            if col_idx in raw_to_template.values():
                continue
            
            # Replace cell references: A4->A5, Z4->Z5, AQ4->AQ5, dll
            formula_new = re.sub(r'([A-Z]+)4\b', r'\g<1>' + str(excel_row), formula_template)
            ws.cell(excel_row, col_idx).value = formula_new
    
    # === UPDATE SUBTOTAL FORMULAS IN ROW 1 (DYNAMIC: cari kolom dengan formula SUBTOTAL) ===
    last_data_row = len(df_raw) + 3  # Row terakhir data
    # Cari kolom yang punya formula SUBTOTAL di row 1 (biasanya kolom kalkulasi)
    for col_idx in range(1, ws.max_column + 1):
        cell_val = ws.cell(1, col_idx).value
        if cell_val and isinstance(cell_val, str) and "SUBTOTAL" in cell_val.upper():
            col_letter = get_column_letter(col_idx)
            ws.cell(1, col_idx).value = f"=SUBTOTAL(9,{col_letter}4:{col_letter}{last_data_row})"
    
    print(f"     ✅ Data Pelanggan updated: {len(df_raw)} baris")

# ==============================================================================
# UPDATE SHEET DATA OPT
# ==============================================================================
def update_sheet_opt(wb, bulan_index):
    """Update sheet Data OPT dengan mapping kolom anti-typo dan formula dinamis"""
    print(f"   > Updating sheet 'Data OPT'...")
    
    if "Data OPT" not in wb.sheetnames:
        print(f"     ⚠️ Sheet 'Data OPT' tidak ditemukan!")
        return
    
    ws = wb["Data OPT"]
    
    # Baca header template dari Row 3
    template_headers = []
    for col in range(1, 150):  # Scan sampai kolom 150
        val = ws.cell(3, col).value
        if val:
            template_headers.append((col, str(val).strip()))
    
    # Baca data baru dari file raw
    try:
        df_raw = pd.read_excel(CONFIG["FILE_OPT"], header=0)
    except FileNotFoundError:
        print(f"     ⚠️ File tidak ditemukan: {CONFIG['FILE_OPT']}")
        return
    except Exception as e:
        print(f"     ❌ Error membaca file OPT: {e}")
        return
    
    # Mapping kolom raw ke template (anti-typo)
    raw_to_template = {}
    for raw_col in df_raw.columns:
        raw_clean = clean_header(raw_col)
        for template_col_idx, template_col_name in template_headers:
            template_clean = clean_header(template_col_name)
            if raw_clean == template_clean:
                raw_to_template[raw_col] = template_col_idx
                break
    
    # === SPECIAL MAPPINGS (kolom yang beda nama antara template vs raw) ===
    # hargaInstallasi (template 2L) ← hargaInstalasi (raw 1L)
    if 'hargaInstalasi' in df_raw.columns:
        for idx, (col_idx, col_name) in enumerate(template_headers):
            if clean_header(col_name) == clean_header('hargaInstallasi'):
                raw_to_template['hargaInstalasi'] = col_idx
                break
    
    # === SORT BY kodeMasterProduk (CUSTOM ORDER dari mentor) ===
    kode_produk_col = None
    for raw_col in df_raw.columns:
        if clean_header(raw_col) == clean_header('kodeMasterProduk'):
            kode_produk_col = raw_col
            break
    
    if kode_produk_col:
        df_raw = custom_sort_by_kode_produk(df_raw, kode_produk_col)
    
    # === UPDATE BULAN BERJALAN (DINAMIS: scan row 1-5) ===
    # Cari label "Bulan Berjalan" di row manapun (1-5), tulis nilai di kolom sebelahnya di row yang sama
    col_label = None
    row_label = None
    for row_idx in range(1, 6):  # Scan row 1-5
        for col_idx in range(1, 150):
            cell_val = ws.cell(row_idx, col_idx).value
            if cell_val and "bulan berjalan" in str(cell_val).lower():
                col_label = col_idx
                row_label = row_idx
                break
        if col_label:
            break
    
    if col_label and row_label:
        # Tulis nilai di kolom sebelah kanan (col_label + 1) di row yang sama
        col_value = col_label + 1
        ws.cell(row_label, col_value).value = bulan_index
        print(f"     ✅ Update Bulan Berjalan (OPT): Row {row_label}, {get_column_letter(col_label)} (label) → {get_column_letter(col_value)} (nilai={bulan_index})")
    else:
        # Fallback ke hardcoded row 2, col 84
        print("     ⚠️  Warning: Kolom 'Bulan Berjalan' tidak ditemukan, gunakan fallback row 2, col 84")
        ws.cell(2, 84).value = bulan_index
    
    # === COPY TEMPLATE ROW 4 FORMULAS ===
    # Simpan formula dari row 4 template untuk di-copy ke row baru
    import re
    template_formulas = {}
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(4, col)
        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
            template_formulas[col] = cell.value
    
    # === HAPUS DATA LAMA (dari row 4 ke bawah) ===
    if ws.max_row > 3:
        ws.delete_rows(4, ws.max_row - 3)
    
    # === TULIS DATA BARU (dari row 4) ===
    for row_idx, row_data in df_raw.iterrows():
        excel_row = row_idx + 4  # Start dari row 4
        
        # 1. TULIS DATA RAW DULU (prioritas tertinggi)
        for raw_col, template_col_idx in raw_to_template.items():
            val = row_data[raw_col]
            if pd.notna(val):
                # Clean invalid date strings (--/--/--, --, dll) untuk mencegah #VALUE!
                str_val = str(val).strip()
                if str_val in ['--/--/--', '--', '---', 'nan', 'NaT', 'None']:
                    # Skip: biarkan cell kosong
                    continue
                ws.cell(excel_row, template_col_idx).value = val
        
        # 2. Copy formulas dari template HANYA untuk kolom yang TIDAK ada di raw data
        for col_idx, formula_template in template_formulas.items():
            # Skip kalau kolom ini sudah ada data raw
            if col_idx in raw_to_template.values():
                continue
            
            # Replace cell references: A4->A5, J4->J5, CC4->CC5, BZ4->BZ5, dll
            formula_new = re.sub(r'([A-Z]+)4\b', r'\g<1>' + str(excel_row), formula_template)
            ws.cell(excel_row, col_idx).value = formula_new
    
    # === UPDATE SUBTOTAL FORMULAS IN ROW 1 (DYNAMIC: cari kolom dengan formula SUBTOTAL) ===
    last_data_row = len(df_raw) + 3  # Row terakhir data
    # Cari kolom yang punya formula SUBTOTAL di row 1
    for col_idx in range(1, ws.max_column + 1):
        cell_val = ws.cell(1, col_idx).value
        if cell_val and isinstance(cell_val, str) and "SUBTOTAL" in cell_val.upper():
            col_letter = get_column_letter(col_idx)
            ws.cell(1, col_idx).value = f"=SUBTOTAL(9,{col_letter}4:{col_letter}{last_data_row})"
    
    print(f"     ✅ Data OPT updated: {len(df_raw)} baris")

# ==============================================================================
# UPDATE RUMUS SUMIF DINAMIS DI DASHBOARD
# ==============================================================================
def update_dashboard_sumif_formulas(wb):
    """Update rumus SUMIF di Dashboard agar range dinamis (ubah statis range jadi full column)"""
    print(f"   > Updating Dashboard SUMIF formulas (dynamic ranges)...")
    
    import re
    ws_dash = wb["Dashboard"]
    
    # Pattern untuk detect & replace range statis jadi dinamis
    # Contoh: $M$4:$M$12429 → $M:$M
    pattern_range = re.compile(r'\$([A-Z]+)\$\d+:\$([A-Z]+)\$\d+')
    
    # Scan semua cell di Dashboard
    for row in ws_dash.iter_rows(min_row=3, max_row=ws_dash.max_row):
        for cell in row:
            # Skip merged cells
            if hasattr(cell, '__class__') and cell.__class__.__name__ == 'MergedCell':
                continue
            
            # Cek apakah cell punya formula
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                original_formula = cell.value
                
                # Replace semua range statis jadi full column
                # $M$4:$M$12429 → $M:$M
                updated_formula = pattern_range.sub(r'$\1:$\2', original_formula)
                
                # Tulis balik kalau ada perubahan
                if updated_formula != original_formula:
                    try:
                        cell.value = updated_formula
                    except:
                        pass  # Skip jika error (merged cell, dll)
    
    print(f"     ✅ Dashboard SUMIF formulas updated: statis range → full column dynamic!")

if __name__ == "__main__":
    try:
        m1, m2, raw, kode_nama_map = extract_data()
        final = transform_data(m1, m2, raw, kode_nama_map)
        load_data(final)
    except FileNotFoundError as e:
        print(f"\n❌ FILE ERROR: {e}")
        print("💡 Pastikan semua file input ada di folder yang sama dengan script ini.")
    except ValueError as e:
        print(f"\n❌ DATA ERROR: {e}")
        print("💡 Cek struktur file Excel (nama sheet, kolom wajib, dll).")
    except Exception as e:
        print(f"\n❌ UNEXPECTED ERROR: {e}")
        import traceback
        traceback.print_exc()